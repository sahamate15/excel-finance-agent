"""Generate ``data/sample_workbook.xlsx`` with the demo Revenue, Assets, and
CashFlows sheets used in the README examples.

Run with: ``python generate_sample_workbook.py``
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUTPUT_PATH = Path(__file__).resolve().parent / "data" / "sample_workbook.xlsx"

_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_INPUT_FONT = Font(color="0000FF", name="Arial", size=11)  # blue = hardcoded input
_BODY_FONT = Font(name="Arial", size=11)
_CURRENCY_FMT = "$#,##0;($#,##0);-"
_PERCENT_FMT = "0.0%"
_CENTER = Alignment(horizontal="center", vertical="center")


def _write_headers(ws, headers: list[str], row: int = 1) -> None:
    """Write a styled header row starting at column 1."""
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER


def _autofit(ws, n_cols: int) -> None:
    """Set a sensible default width for the first *n_cols* columns."""
    for c in range(1, n_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18


def build_revenue_sheet(wb: Workbook) -> None:
    """Populate the Revenue sheet (years, revenue, cost, empty EBITDA)."""
    ws = wb.create_sheet("Revenue")
    _write_headers(ws, ["Year", "Revenue ($)", "Cost ($)", "EBITDA ($)"])
    years = [2020, 2021, 2022, 2023, 2024]
    revenue = [50_000, 62_000, 71_000, 85_000, 102_000]
    cost = [30_000, 37_000, 42_000, 49_000, 58_000]
    for i, (y, r, c) in enumerate(zip(years, revenue, cost), start=2):
        ws.cell(row=i, column=1, value=y).number_format = "0"
        rc = ws.cell(row=i, column=2, value=r)
        rc.number_format = _CURRENCY_FMT
        rc.font = _INPUT_FONT
        cc = ws.cell(row=i, column=3, value=c)
        cc.number_format = _CURRENCY_FMT
        cc.font = _INPUT_FONT
        # Column D (EBITDA) deliberately left empty for the agent to fill.
    _autofit(ws, 4)


def build_assets_sheet(wb: Workbook) -> None:
    """Populate the Assets sheet (asset, purchase value, depreciation rate)."""
    ws = wb.create_sheet("Assets")
    _write_headers(ws, ["Asset", "Purchase Value ($)", "Rate"])
    rows = [
        ("Machinery", 2_000_000, 0.25),
        ("Computers", 500_000, 0.40),
    ]
    for i, (name, value, rate) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=name).font = _BODY_FONT
        v = ws.cell(row=i, column=2, value=value)
        v.font = _INPUT_FONT
        v.number_format = _CURRENCY_FMT
        r = ws.cell(row=i, column=3, value=rate)
        r.font = _INPUT_FONT
        r.number_format = _PERCENT_FMT
    _autofit(ws, 3)


def build_cashflows_sheet(wb: Workbook) -> None:
    """Populate the CashFlows sheet (year 0..5, cashflow values)."""
    ws = wb.create_sheet("CashFlows")
    _write_headers(ws, ["Year", "Cash Flow ($)"])
    flows = [-1_000_000, 250_000, 300_000, 350_000, 400_000, 450_000]
    for i, cf in enumerate(flows, start=2):
        ws.cell(row=i, column=1, value=i - 2).number_format = "0"
        c = ws.cell(row=i, column=2, value=cf)
        c.font = _INPUT_FONT
        c.number_format = _CURRENCY_FMT
    _autofit(ws, 2)


def main() -> None:
    """Build the workbook and save it to :data:`OUTPUT_PATH`."""
    wb = Workbook()
    # Remove the default empty sheet that openpyxl creates.
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    build_revenue_sheet(wb)
    build_assets_sheet(wb)
    build_cashflows_sheet(wb)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"✓ Sample workbook written to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
