"""Master orchestrator: takes a free-text instruction, decides what to do,
delegates to the AI engine and writer, and returns a structured result.

This is the single entry point used by both ``main.py`` (CLI) and
``app.py`` (Streamlit), so its return contract — a fully-keyed result dict —
must remain stable.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from agents.ai_engine import (
    _heuristic_table_config,
    _keyword_classify,
    clarify_input,
    detect_task_type,
    extract_table_config,
    text_to_formula,
)
from agents.excel_writer import (
    create_table,
    find_next_empty_cell,
    preview_table,
    write_formula,
)
from config import CONFIG
from finance.formula_map import get_fallback_formula
from utils.audit import (
    record_clarification_requested,
    record_dry_run_preview,
    record_formula_generated,
    record_formula_rejected,
    record_formula_write,
    record_formula_write_failed,
    record_input_rejected,
    record_instruction_submitted,
    record_table_build_failed,
    record_table_built,
    record_table_config_extracted,
    record_task_type_detected,
)
from utils.logger import get_logger
from utils.validators import (
    sanitize_instruction,
    validate_cell_reference,
    validate_excel_formula,
    validate_table_config,
)

logger = get_logger(__name__)


def _empty_result() -> dict[str, Any]:
    """Return a fresh result skeleton with all keys set to safe defaults."""
    return {
        "success": False,
        "task_type": None,
        "formula": None,
        "cell_written": None,
        "table_created": False,
        "dry_run": False,
        "preview": None,
        "source": None,
        "message": "",
        "error": None,
    }


def execute_excel_task(
    instruction: str,
    filepath: str | Path,
    sheet_name: str = "Sheet1",
    cell: str | None = None,
    *,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Run the full pipeline for one user instruction.

    Args:
        instruction: Plain-English finance instruction.
        filepath:    Path to the target .xlsx file.
        sheet_name:  Sheet to operate on (created if missing).
        cell:        Optional A1-style target cell. If omitted for formula
            tasks, the writer picks the next empty cell automatically.
        dry_run:     When True, returns a structured preview of every cell
            and formula that *would* be written. The workbook on disk is not
            touched and no audit event is emitted. Use a follow-up call with
            ``dry_run=False`` to commit.

    Returns:
        A dict with keys:
        ``success``, ``task_type``, ``formula``, ``cell_written``,
        ``table_created``, ``dry_run``, ``preview``, ``message``, ``error``.
    """
    result = _empty_result()
    result["dry_run"] = dry_run
    instruction = sanitize_instruction(instruction)
    logger.info(
        "Received instruction (%d chars, strict_mode=%s, dry_run=%s)",
        len(instruction), CONFIG.strict_mode, dry_run,
    )

    if not instruction:
        record_input_rejected(reason="empty_instruction", field="instruction")
        result["error"] = "Empty instruction"
        result["message"] = "Please provide a non-empty instruction."
        return result

    # Audit: instruction text. Compliance reviews this. Cell values never go here.
    record_instruction_submitted(instruction=instruction)

    # 1. Classify. In strict mode the LLM is bypassed entirely.
    try:
        if CONFIG.strict_mode:
            task = _keyword_classify(instruction)
            classifier_source = "offline_classifier"
        else:
            task = detect_task_type(instruction)
            classifier_source = "llm"
    except Exception as exc:  # noqa: BLE001 — we want a graceful failure
        logger.exception("Task classification failed")
        result["error"] = f"Classification failed: {exc}"
        result["message"] = "Could not classify the instruction."
        return result

    result["task_type"] = task["type"]
    record_task_type_detected(
        task_type=task["type"],
        source=classifier_source,
        requires_clarification=bool(task.get("requires_clarification")),
    )
    logger.info(
        "Task classified type=%s complexity=%s requires_clarification=%s",
        task["type"], task["complexity"], task["requires_clarification"],
    )

    # 2. Clarification gate
    if task.get("requires_clarification"):
        if CONFIG.strict_mode:
            # The keyword classifier never asks for clarification, but be defensive.
            result["message"] = (
                "Strict mode is enabled and the offline classifier flagged this "
                "instruction as ambiguous. Rephrase with explicit cell references "
                "and numeric inputs, or disable STRICT_MODE in .env to use the LLM."
            )
            result["error"] = "strict_mode_needs_clarification"
            return result
        question = task.get("clarification_question") or clarify_input(instruction)
        record_clarification_requested(question=question)
        result["message"] = question
        result["error"] = "needs_clarification"
        return result

    # 3. Dispatch by type
    if task["type"] == "formula":
        return _handle_formula(instruction, filepath, sheet_name, cell, result, dry_run=dry_run)
    if task["type"] == "table":
        return _handle_table(instruction, filepath, sheet_name, result, dry_run=dry_run)
    if task["type"] == "chart":
        result["message"] = (
            "Chart generation is not implemented yet. Try a formula or "
            "table instruction such as 'create a 5-year revenue projection'."
        )
        result["error"] = "unsupported_task_type"
        return result

    result["message"] = f"Unknown task type: {task['type']}"
    result["error"] = "unknown_task_type"
    return result


# ──────────────────────────────────────────────────────────────────────
# Sub-handlers
# ──────────────────────────────────────────────────────────────────────


def _handle_formula(
    instruction: str,
    filepath: str | Path,
    sheet_name: str,
    cell: str | None,
    result: dict[str, Any],
    *,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Resolve and (unless dry_run) write a single-cell formula."""
    formula: str | None = None
    source: str = "offline_fallback"

    # Strict mode: only the deterministic offline map is allowed.
    if CONFIG.strict_mode:
        formula = get_fallback_formula(instruction)
        if formula is None:
            result["error"] = "strict_mode_unhandled_formula"
            result["message"] = (
                "Strict mode is enabled and the offline formula map cannot "
                "resolve this instruction. Disable STRICT_MODE in .env to use "
                "the LLM, or rephrase using known patterns (growth rate, IRR, "
                "NPV, EMI, CAGR, average, sum, etc.)."
            )
            return result
        source = "offline_fallback"
    else:
        # Fast path: hardcoded fallback.
        if CONFIG.fallback_first:
            formula = get_fallback_formula(instruction)
            if formula:
                source = "offline_fallback"
                logger.info("Resolved via fallback map")

        # Slow path: LLM.
        if formula is None:
            try:
                formula = text_to_formula(instruction, context={"target_cell": cell})
                source = "llm"
            except (ValueError, RuntimeError) as exc:
                # Last-ditch attempt to use the fallback even if FALLBACK_FIRST=False.
                formula = get_fallback_formula(instruction)
                if formula is None:
                    result["error"] = type(exc).__name__
                    result["message"] = (
                        "Could not generate a formula. Try rephrasing or "
                        "specifying cell references explicitly."
                    )
                    return result
                source = "offline_fallback"
                logger.warning("LLM failed (%s); used fallback", type(exc).__name__)

    ok, err = validate_excel_formula(formula)
    if not ok:
        record_formula_rejected(formula=formula, reason=err)
        result["error"] = f"Invalid formula: {err}"
        result["message"] = result["error"]
        return result

    target_cell = cell
    if target_cell is None or not validate_cell_reference(target_cell):
        target_cell = _next_empty_for_file(filepath, sheet_name)

    record_formula_generated(
        formula=formula, target_cell=target_cell, sheet=sheet_name, source=source,
    )

    if dry_run:
        record_dry_run_preview(sheet=sheet_name, cells=1, target_cell=target_cell)
        result["success"] = True
        result["formula"] = formula
        result["cell_written"] = None
        result["preview"] = [{
            "sheet": sheet_name,
            "cell": target_cell,
            "value": formula,
        }]
        result["message"] = (
            f"[dry-run] Would write {formula} to {sheet_name}!{target_cell}. "
            "Re-run with dry_run=False to commit."
        )
        return result

    try:
        write_formula(filepath, sheet_name, target_cell, formula)
    except (ValueError, InvalidFileException, OSError) as exc:
        logger.exception("write_formula failed")
        record_formula_write_failed(
            file=filepath, sheet=sheet_name, cell=target_cell, error=str(exc),
        )
        result["error"] = str(exc)
        result["message"] = f"Failed to write formula: {exc}"
        return result

    record_formula_write(
        file=filepath,
        sheet=sheet_name,
        cell=target_cell,
        formula=formula,
        source=source,
    )

    result["success"] = True
    result["formula"] = formula
    result["cell_written"] = target_cell
    result["source"] = source
    result["message"] = f"Wrote {formula} to {sheet_name}!{target_cell}"
    return result


def _handle_table(
    instruction: str,
    filepath: str | Path,
    sheet_name: str,
    result: dict[str, Any],
    *,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Build a multi-row table from the instruction (or preview if dry_run)."""
    table_config: dict | None = None
    source: str = "offline_fallback"

    if CONFIG.strict_mode:
        table_config = _heuristic_table_config(instruction, strict=True)
        if table_config is None:
            result["error"] = "strict_mode_unhandled_table"
            result["message"] = (
                "Strict mode is enabled and the offline parser cannot extract "
                "a table config from this instruction. Rephrase to mention "
                "depreciation, amortization, or projection explicitly, or "
                "disable STRICT_MODE in .env to use the LLM."
            )
            return result
        source = "offline_fallback"
    else:
        try:
            table_config = extract_table_config(instruction)
            source = "llm"
        except (ValueError, RuntimeError) as exc:
            logger.warning("extract_table_config LLM path failed (%s); using heuristic", type(exc).__name__)
            table_config = _heuristic_table_config(instruction, strict=False)
            source = "offline_fallback"

    logger.info("Table type: %s (source=%s)", table_config.get("type"), source)
    if "type" not in table_config:
        result["error"] = "table_config missing 'type'"
        result["message"] = "Could not infer table type from instruction."
        return result

    record_table_config_extracted(
        table_type=str(table_config.get("type")), source=source,
    )

    # Validate before either previewing or writing — same gate for both paths.
    ok, err = validate_table_config(table_config)
    if not ok:
        record_input_rejected(reason=err, field="table_config")
        result["error"] = "invalid_table_config"
        result["message"] = f"Invalid table config: {err}"
        return result

    if dry_run:
        try:
            preview = preview_table(table_config, sheet_name=sheet_name)
        except ValueError as exc:
            result["error"] = "preview_failed"
            result["message"] = f"Failed to build preview: {exc}"
            return result
        record_dry_run_preview(sheet=sheet_name, cells=len(preview))
        result["success"] = True
        result["preview"] = preview
        result["table_created"] = False
        result["message"] = (
            f"[dry-run] Would build a {table_config.get('type')} table on "
            f"sheet {sheet_name!r} ({len(preview)} cells). "
            "Re-run with dry_run=False to commit."
        )
        result["formula"] = None
        return result

    try:
        create_table(filepath, sheet_name, table_config)
    except (ValueError, InvalidFileException, OSError, KeyError) as exc:
        logger.exception("create_table failed")
        record_table_build_failed(
            file=filepath, sheet=sheet_name,
            table_type=str(table_config.get("type")),
            error=str(exc),
        )
        result["error"] = str(exc)
        result["message"] = f"Failed to create table: {exc}"
        return result

    record_table_built(
        file=filepath,
        sheet=sheet_name,
        table_type=str(table_config.get("type")),
        source=source,
    )

    result["success"] = True
    result["table_created"] = True
    result["source"] = source
    result["message"] = (
        f"Created {table_config.get('type')} table on sheet {sheet_name!r}."
    )
    result["formula"] = None
    return result


def _next_empty_for_file(filepath: str | Path, sheet_name: str) -> str:
    """Open the workbook just to ask the writer for the next empty cell."""
    path = Path(filepath)
    if not path.exists():
        return "A1"
    try:
        wb = load_workbook(path)
    except InvalidFileException:
        return "A1"
    if sheet_name not in wb.sheetnames:
        return "A1"
    return find_next_empty_cell(wb[sheet_name])
