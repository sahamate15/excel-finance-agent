"""All openpyxl-facing logic — write a single formula or build a finance table.

The writer never calculates results in Python; every numeric output is an
Excel formula so the workbook stays dynamic when assumptions change. Tables
ship with light, professional styling (blue header, alternating rows,
borders, currency / percent formats).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from utils.logger import get_logger
from utils.validators import (
    validate_cell_reference,
    validate_excel_formula,
    validate_table_config,
)

logger = get_logger(__name__)

# ──────────────────────────────────────────────────────────────────────
# Styling constants
# ──────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
_ROW_FILL_LIGHT = PatternFill("solid", fgColor="F2F2F2")
_ROW_FILL_WHITE = PatternFill("solid", fgColor="FFFFFF")
_HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")  # light yellow
_BODY_FONT = Font(name="Arial", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")

_CURRENCY_FMT = "#,##0.00;(#,##0.00);-"
_PERCENT_FMT = "0.00%"


# ──────────────────────────────────────────────────────────────────────
# Workbook open / save
# ──────────────────────────────────────────────────────────────────────


def _load_or_create(filepath: Path) -> Workbook:
    """Open the workbook if it exists, otherwise create a fresh one.

    Args:
        filepath: Path to the .xlsx file.

    Returns:
        An openpyxl ``Workbook``.

    Raises:
        InvalidFileException: If the file exists but cannot be parsed.
    """
    filepath = Path(filepath)
    if filepath.exists():
        try:
            return load_workbook(filepath)
        except InvalidFileException:
            logger.exception("Cannot open workbook at %s", filepath)
            raise
    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Drop the auto-created Sheet so the caller can choose a name.
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]
    return wb


def _ensure_sheet(wb: Workbook, sheet_name: str) -> Worksheet:
    """Return the named sheet, creating it if missing."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


# ──────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────


def write_formula(filepath: str | Path, sheet_name: str, cell: str, formula: str) -> bool:
    """Write *formula* to *cell* of *sheet_name* in *filepath*.

    Args:
        filepath:   Path to the .xlsx file (created if missing).
        sheet_name: Sheet to write into (created if missing).
        cell:       A1-style cell reference (e.g. ``"C5"``).
        formula:    Formula string starting with ``=``.

    Returns:
        ``True`` on success.

    Raises:
        ValueError: If *cell* or *formula* is invalid.
        InvalidFileException: If the existing file is corrupt.
    """
    if not validate_cell_reference(cell):
        raise ValueError(f"Invalid cell reference: {cell!r}")
    ok, err = validate_excel_formula(formula)
    if not ok:
        raise ValueError(f"Invalid formula: {err}")

    path = Path(filepath)
    wb = _load_or_create(path)
    ws = _ensure_sheet(wb, sheet_name)

    target = ws[cell.upper()]
    target.value = formula
    target.fill = _HIGHLIGHT_FILL
    target.font = _BODY_FONT

    wb.save(path)
    logger.info("Wrote formula %s to %s!%s in %s", formula, sheet_name, cell, path.name)
    return True


def find_next_empty_cell(ws: Worksheet, preferred_col: str | int | None = None) -> str:
    """Locate the next empty cell to write into.

    Args:
        ws:            Active worksheet.
        preferred_col: If provided, find the next empty row in this column
            (letter or 1-based int). Otherwise pick the first column that is
            entirely empty and return its row 2.

    Returns:
        A cell address string such as ``"C5"``.
    """
    if preferred_col is not None:
        col_letter = (
            preferred_col.upper() if isinstance(preferred_col, str)
            else get_column_letter(int(preferred_col))
        )
        col_idx = ws[f"{col_letter}1"].column
        # Start from row 2 (assume row 1 is header).
        for row in range(2, ws.max_row + 2):
            if ws.cell(row=row, column=col_idx).value in (None, ""):
                return f"{col_letter}{row}"
        return f"{col_letter}{ws.max_row + 1}"

    # No preference: scan columns left-to-right for the first entirely empty one.
    for col_idx in range(1, ws.max_column + 2):
        column_values = [ws.cell(row=r, column=col_idx).value for r in range(1, ws.max_row + 1)]
        if all(v in (None, "") for v in column_values):
            return f"{get_column_letter(col_idx)}2"
    return f"{get_column_letter(ws.max_column + 1)}2"


def create_table(filepath: str | Path, sheet_name: str, table_config: dict) -> bool:
    """Build one of the supported finance tables in the named sheet.

    Supported ``table_config['type']`` values: ``"depreciation"``,
    ``"amortization"``, ``"projection"``. Configs are validated against
    finance-domain bounds before any disk write — see
    :func:`utils.validators.validate_table_config`.

    Args:
        filepath:     Path to the .xlsx file (created if missing).
        sheet_name:   Sheet to write into (created if missing).
        table_config: Dict matching one of the schemas in
            :mod:`agents.ai_engine` (see ``TABLE_CONFIG_SYSTEM_PROMPT``).

    Returns:
        ``True`` on success.

    Raises:
        ValueError: If the ``type`` field is missing/unsupported, or if any
            field is outside its allowed range.
    """
    ok, err = validate_table_config(table_config)
    if not ok:
        raise ValueError(f"Invalid table config: {err}")

    table_type = table_config.get("type")
    path = Path(filepath)
    wb = _load_or_create(path)
    ws = _ensure_sheet(wb, sheet_name)

    if table_type == "depreciation":
        _build_depreciation(ws, table_config)
    elif table_type == "amortization":
        _build_amortization(ws, table_config)
    elif table_type == "projection":
        _build_projection(ws, table_config)
    else:
        raise ValueError(f"Unsupported table type: {table_type!r}")

    wb.save(path)
    logger.info("Created %s table on %s!%s", table_type, sheet_name, ws.title)
    return True


# ──────────────────────────────────────────────────────────────────────
# Table builders
# ──────────────────────────────────────────────────────────────────────


def _build_depreciation(ws: Worksheet, cfg: dict) -> None:
    """Lay out a 4-column depreciation schedule.

    Supports two methods via ``cfg["method"]``:
    - ``"wdv"`` (default): written-down value. Each year's depreciation is
      ``Opening * Rate``; closing = opening − depreciation.
    - ``"straight_line"``: per-year depreciation is constant at
      ``(asset_value − salvage_value) / years``. ``rate`` is not consulted.

    Optional ``cfg["salvage_value"]`` (default 0) is used by straight-line
    and is also displayed as an assumption block for WDV reference.

    All per-row depreciation amounts are wrapped in ``ROUND(..., 2)`` so
    accumulated totals tie to two decimal places.
    """
    asset_value = float(cfg["asset_value"])
    years = int(cfg["years"])
    start_row = int(cfg.get("start_row", 1))
    start_col = int(cfg.get("start_col", 1))
    method = cfg.get("method", "wdv")
    salvage_value = float(cfg.get("salvage_value", 0))

    headers = ["Year", "Opening Value", "Depreciation", "Closing Value"]
    _write_headers(ws, headers, start_row, start_col)

    # Assumption block placed two columns to the right of the table.
    assum_col = start_col + len(headers) + 1
    ws.cell(row=start_row, column=assum_col, value="Inputs").font = Font(bold=True, name="Arial")

    if method == "wdv":
        rate = float(cfg["rate"])
        ws.cell(row=start_row + 1, column=assum_col, value="Rate")
        rate_cell = ws.cell(row=start_row + 1, column=assum_col + 1, value=rate)
        rate_cell.number_format = _PERCENT_FMT
        rate_cell.font = Font(color="0000FF", name="Arial")  # blue = input
        rate_ref = f"${get_column_letter(assum_col + 1)}${start_row + 1}"
    else:
        # Straight-line: constant per-year depreciation = (cost - salvage) / years.
        ws.cell(row=start_row + 1, column=assum_col, value="Salvage Value")
        sv_cell = ws.cell(row=start_row + 1, column=assum_col + 1, value=salvage_value)
        sv_cell.number_format = _CURRENCY_FMT
        sv_cell.font = Font(color="0000FF", name="Arial")
        ws.cell(row=start_row + 2, column=assum_col, value="Useful Life (years)")
        life_cell = ws.cell(row=start_row + 2, column=assum_col + 1, value=years)
        life_cell.font = Font(color="0000FF", name="Arial")
        sv_ref = f"${get_column_letter(assum_col + 1)}${start_row + 1}"
        life_ref = f"${get_column_letter(assum_col + 1)}${start_row + 2}"

    for i in range(years):
        row = start_row + 1 + i
        year_num = i + 1
        opening_col = get_column_letter(start_col + 1)
        deprec_col = get_column_letter(start_col + 2)
        closing_col = get_column_letter(start_col + 3)

        # Year
        ws.cell(row=row, column=start_col, value=year_num).number_format = "0"

        # Opening value: literal in row 1, otherwise pulls previous closing.
        if i == 0:
            ws.cell(row=row, column=start_col + 1, value=asset_value)
        else:
            prev_closing = f"{closing_col}{row - 1}"
            ws.cell(row=row, column=start_col + 1, value=f"={prev_closing}")

        # Depreciation: WDV uses opening * rate, straight-line uses (cost - salvage) / life.
        if method == "wdv":
            deprec_formula = f"=ROUND({opening_col}{row}*{rate_ref},2)"
        else:
            # asset_value is the literal in row 1 of the Opening Value column.
            cost_ref = f"${opening_col}${start_row + 1}"
            deprec_formula = f"=ROUND(({cost_ref}-{sv_ref})/{life_ref},2)"
        ws.cell(row=row, column=start_col + 2, value=deprec_formula)

        # Closing = Opening − Depreciation
        ws.cell(
            row=row,
            column=start_col + 3,
            value=f"={opening_col}{row}-{deprec_col}{row}",
        )

    end_row = start_row + years
    end_col = start_col + len(headers) - 1
    _apply_table_style(ws, start_row, end_row, start_col, end_col)
    # Currency format on the value columns (cols 2..4)
    for c in range(start_col + 1, start_col + 4):
        for r in range(start_row + 1, end_row + 1):
            ws.cell(row=r, column=c).number_format = _CURRENCY_FMT


def _build_amortization(ws: Worksheet, cfg: dict) -> None:
    """Lay out a monthly loan amortization schedule using PMT/IPMT/PPMT."""
    principal = float(cfg["principal"])
    annual_rate = float(cfg["annual_rate"])
    tenure_months = int(cfg["tenure_months"])
    start_row = int(cfg.get("start_row", 1))
    start_col = int(cfg.get("start_col", 1))

    headers = [
        "Month", "Opening Balance", "EMI", "Interest", "Principal", "Closing Balance",
    ]
    _write_headers(ws, headers, start_row, start_col)

    # Assumption block to the right of the table.
    assum_col = start_col + len(headers) + 1
    ws.cell(row=start_row, column=assum_col, value="Inputs").font = Font(bold=True, name="Arial")
    ws.cell(row=start_row + 1, column=assum_col, value="Principal")
    p_cell = ws.cell(row=start_row + 1, column=assum_col + 1, value=principal)
    p_cell.font = Font(color="0000FF", name="Arial")
    p_cell.number_format = _CURRENCY_FMT
    ws.cell(row=start_row + 2, column=assum_col, value="Annual Rate")
    r_cell = ws.cell(row=start_row + 2, column=assum_col + 1, value=annual_rate)
    r_cell.font = Font(color="0000FF", name="Arial")
    r_cell.number_format = _PERCENT_FMT
    ws.cell(row=start_row + 3, column=assum_col, value="Tenure (months)")
    t_cell = ws.cell(row=start_row + 3, column=assum_col + 1, value=tenure_months)
    t_cell.font = Font(color="0000FF", name="Arial")

    p_ref = f"${get_column_letter(assum_col + 1)}${start_row + 1}"
    r_ref = f"${get_column_letter(assum_col + 1)}${start_row + 2}"
    t_ref = f"${get_column_letter(assum_col + 1)}${start_row + 3}"

    month_col = get_column_letter(start_col)
    open_col = get_column_letter(start_col + 1)
    close_col = get_column_letter(start_col + 5)

    for i in range(tenure_months):
        row = start_row + 1 + i
        month_num = i + 1
        ws.cell(row=row, column=start_col, value=month_num).number_format = "0"

        # Opening balance
        if i == 0:
            ws.cell(row=row, column=start_col + 1, value=f"={p_ref}")
        else:
            ws.cell(
                row=row,
                column=start_col + 1,
                value=f"={close_col}{row - 1}",
            )

        # EMI (constant) — positive number, hence -PMT
        ws.cell(
            row=row,
            column=start_col + 2,
            value=f"=-PMT({r_ref}/12,{t_ref},{p_ref})",
        )
        # Interest portion this month — period reads from the Month column.
        ws.cell(
            row=row,
            column=start_col + 3,
            value=f"=-IPMT({r_ref}/12,{month_col}{row},{t_ref},{p_ref})",
        )
        # Principal portion this month
        ws.cell(
            row=row,
            column=start_col + 4,
            value=f"=-PPMT({r_ref}/12,{month_col}{row},{t_ref},{p_ref})",
        )
        # Closing balance = Opening - Principal portion
        ws.cell(
            row=row,
            column=start_col + 5,
            value=f"={open_col}{row}-{get_column_letter(start_col + 4)}{row}",
        )

    end_row = start_row + tenure_months
    end_col = start_col + len(headers) - 1
    _apply_table_style(ws, start_row, end_row, start_col, end_col)
    for c in range(start_col + 1, end_col + 1):
        for r in range(start_row + 1, end_row + 1):
            ws.cell(row=r, column=c).number_format = _CURRENCY_FMT


def _build_projection(ws: Worksheet, cfg: dict) -> None:
    """Lay out a multi-year revenue / costs / EBITDA projection.

    Supports two growth methods via ``cfg["growth_method"]``:
    - ``"compound"`` (default): Year N = previous year * (1 + growth_rate).
      Equivalent to ``base * (1 + r)^(N-1)``. CAGR-style.
    - ``"linear"``: Year N = base * (1 + growth_rate * (N-1)). Same absolute
      increment each year. Sometimes called "flat growth" in IB context.
    """
    base_revenue = float(cfg["base_revenue"])
    growth_rate = float(cfg["growth_rate"])
    cost_ratio = float(cfg.get("cost_ratio", 0.6))
    years = int(cfg.get("years", 5))
    start_row = int(cfg.get("start_row", 1))
    start_col = int(cfg.get("start_col", 1))
    growth_method = cfg.get("growth_method", "compound")

    headers = ["Year", "Revenue", "Costs", "EBITDA", "EBITDA Margin %"]
    _write_headers(ws, headers, start_row, start_col)

    # Assumption block to the right.
    assum_col = start_col + len(headers) + 1
    ws.cell(row=start_row, column=assum_col, value="Inputs").font = Font(bold=True, name="Arial")
    ws.cell(row=start_row + 1, column=assum_col, value="Base Revenue")
    br_cell = ws.cell(row=start_row + 1, column=assum_col + 1, value=base_revenue)
    br_cell.font = Font(color="0000FF", name="Arial")
    br_cell.number_format = _CURRENCY_FMT
    ws.cell(row=start_row + 2, column=assum_col, value="Growth Rate")
    g_cell = ws.cell(row=start_row + 2, column=assum_col + 1, value=growth_rate)
    g_cell.font = Font(color="0000FF", name="Arial")
    g_cell.number_format = _PERCENT_FMT
    ws.cell(row=start_row + 3, column=assum_col, value="Cost Ratio")
    c_cell = ws.cell(row=start_row + 3, column=assum_col + 1, value=cost_ratio)
    c_cell.font = Font(color="0000FF", name="Arial")
    c_cell.number_format = _PERCENT_FMT

    br_ref = f"${get_column_letter(assum_col + 1)}${start_row + 1}"
    g_ref = f"${get_column_letter(assum_col + 1)}${start_row + 2}"
    cr_ref = f"${get_column_letter(assum_col + 1)}${start_row + 3}"

    rev_col = get_column_letter(start_col + 1)
    cost_col = get_column_letter(start_col + 2)
    ebitda_col = get_column_letter(start_col + 3)

    for i in range(years):
        row = start_row + 1 + i
        year_num = i + 1
        ws.cell(row=row, column=start_col, value=year_num).number_format = "0"

        if i == 0:
            ws.cell(row=row, column=start_col + 1, value=f"={br_ref}")
        elif growth_method == "compound":
            # Compounding off the previous year — geometric growth.
            ws.cell(
                row=row,
                column=start_col + 1,
                value=f"={rev_col}{row - 1}*(1+{g_ref})",
            )
        else:
            # Linear: each year = base * (1 + r * (N-1)). Same absolute
            # increment each year, so manual edits to mid-year revenue do not
            # cascade.
            ws.cell(
                row=row,
                column=start_col + 1,
                value=f"={br_ref}*(1+{g_ref}*{year_num - 1})",
            )
        ws.cell(row=row, column=start_col + 2, value=f"={rev_col}{row}*{cr_ref}")
        ws.cell(
            row=row,
            column=start_col + 3,
            value=f"={rev_col}{row}-{cost_col}{row}",
        )
        ws.cell(
            row=row,
            column=start_col + 4,
            value=f"={ebitda_col}{row}/{rev_col}{row}",
        )

    end_row = start_row + years
    end_col = start_col + len(headers) - 1
    _apply_table_style(ws, start_row, end_row, start_col, end_col)
    for c in range(start_col + 1, start_col + 4):
        for r in range(start_row + 1, end_row + 1):
            ws.cell(row=r, column=c).number_format = _CURRENCY_FMT
    for r in range(start_row + 1, end_row + 1):
        ws.cell(row=r, column=start_col + 4).number_format = _PERCENT_FMT


# ──────────────────────────────────────────────────────────────────────
# Styling helpers
# ──────────────────────────────────────────────────────────────────────


def _write_headers(ws: Worksheet, headers: list[str], start_row: int, start_col: int) -> None:
    """Write a header row with the canonical bold-blue style."""
    for offset, label in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + offset, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _apply_table_style(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """Apply borders, alternating row fills, alignment and column widths.

    Args:
        ws:        Worksheet being styled.
        start_row: Row of the header (1-based).
        end_row:   Last data row (1-based, inclusive).
        start_col: First column index (1-based).
        end_col:   Last column index (1-based, inclusive).
    """
    for row in range(start_row + 1, end_row + 1):
        fill = _ROW_FILL_LIGHT if (row - start_row) % 2 == 0 else _ROW_FILL_WHITE
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = _BODY_FONT
            cell.border = _THIN_BORDER
            cell.alignment = _RIGHT if col > start_col else _CENTER

    # Column widths — estimate from header length + a margin.
    for col in range(start_col, end_col + 1):
        header_val = ws.cell(row=start_row, column=col).value
        width = max(14, len(str(header_val)) + 4) if header_val else 14
        ws.column_dimensions[get_column_letter(col)].width = width


def apply_table_style(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    """Public wrapper around :func:`_apply_table_style` for external callers."""
    _apply_table_style(ws, start_row, end_row, start_col, end_col)


def preview_table(table_config: dict, sheet_name: str = "Preview") -> list[dict[str, Any]]:
    """Build a table in memory and return every populated cell as a preview.

    No file is touched. Used by the dry-run path so a user can see exactly
    what formulas and values would be written before committing.

    Args:
        table_config: Config validated by :func:`validate_table_config`.
        sheet_name:   Sheet name to label preview entries with.

    Returns:
        A list of ``{"sheet": str, "cell": str, "value": Any}`` dicts, in
        worksheet iteration order. Formulas are returned as their string
        form (starting with ``=``); literals are the underlying type.

    Raises:
        ValueError: If the config is invalid (same checks as ``create_table``).
    """
    ok, err = validate_table_config(table_config)
    if not ok:
        raise ValueError(f"Invalid table config: {err}")

    wb = Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        del wb["Sheet"]
    ws = wb.create_sheet(sheet_name)

    table_type = table_config.get("type")
    if table_type == "depreciation":
        _build_depreciation(ws, table_config)
    elif table_type == "amortization":
        _build_amortization(ws, table_config)
    elif table_type == "projection":
        _build_projection(ws, table_config)
    else:
        raise ValueError(f"Unsupported table type: {table_type!r}")

    preview: list[dict[str, Any]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.value == "":
                continue
            preview.append({
                "sheet": sheet_name,
                "cell": cell.coordinate,
                "value": cell.value,
            })
    return preview


__all__ = [
    "write_formula",
    "find_next_empty_cell",
    "create_table",
    "apply_table_style",
    "preview_table",
]


# Suppress unused-import lint when run as a script.
_ = Any
