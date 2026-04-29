"""Pytest suite for the Excel Finance Automation Agent.

Tests that contact Mistral's API are guarded with ``skipif`` based on the
presence of ``MISTRAL_API_KEY``, so the suite is fully runnable offline. The
file-IO tests use ``tmp_path`` so nothing leaks into the repo.
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# Make project root importable when pytest is run from anywhere.
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from agents.excel_writer import (  # noqa: E402
    create_table,
    find_next_empty_cell,
    preview_table,
    write_formula,
)
from finance.formula_map import get_fallback_formula  # noqa: E402
from utils.validators import (  # noqa: E402
    sanitize_instruction,
    validate_cell_reference,
    validate_excel_formula,
    validate_table_config,
)

NEEDS_MISTRAL = pytest.mark.skipif(
    not os.getenv("MISTRAL_API_KEY"),
    reason="MISTRAL_API_KEY not configured — skipping live LLM tests.",
)


# ──────────────────────────────────────────────────────────────────────
# validate_excel_formula
# ──────────────────────────────────────────────────────────────────────


class TestValidateFormula:
    """Lightweight syntactic checks for formula strings."""

    def test_valid_sum(self) -> None:
        ok, err = validate_excel_formula("=SUM(A1:A10)")
        assert ok is True
        assert err == ""

    def test_missing_equals(self) -> None:
        ok, err = validate_excel_formula("SUM(A1:A10)")
        assert ok is False
        assert "must start with =" in err

    def test_unbalanced_parens(self) -> None:
        ok, err = validate_excel_formula("=(A1+B1")
        assert ok is False
        assert "unbalanced parentheses" in err

    def test_empty_body(self) -> None:
        ok, _ = validate_excel_formula("=")
        assert ok is False

    def test_arithmetic_only(self) -> None:
        ok, _ = validate_excel_formula("=A1+B1")
        assert ok is True

    def test_too_long(self) -> None:
        ok, err = validate_excel_formula("=" + "A1+" * 400 + "1")
        assert ok is False
        assert "exceeds" in err

    def test_control_chars_rejected(self) -> None:
        ok, err = validate_excel_formula("=SUM(A1:A10)\n")
        assert ok is False
        assert "forbidden" in err


# ──────────────────────────────────────────────────────────────────────
# validate_cell_reference + sanitize_instruction
# ──────────────────────────────────────────────────────────────────────


class TestValidators:
    """Cell-reference parsing and instruction sanitisation."""

    @pytest.mark.parametrize("ref", ["A1", "BC100", "XFD1048576", "Z9"])
    def test_valid_refs(self, ref: str) -> None:
        assert validate_cell_reference(ref) is True

    @pytest.mark.parametrize("ref", ["1A", "AAAA1", "A0", "", "A-1", "A1.5"])
    def test_invalid_refs(self, ref: str) -> None:
        assert validate_cell_reference(ref) is False

    def test_sanitize_strips_control(self) -> None:
        assert sanitize_instruction("calc IRR\nfor B2:B7") == "calc IRR for B2:B7"

    def test_sanitize_truncates(self) -> None:
        long = "x" * 1000
        assert len(sanitize_instruction(long)) == 500


# ──────────────────────────────────────────────────────────────────────
# Fallback formula map
# ──────────────────────────────────────────────────────────────────────


class TestFallbackFormula:
    """Hardcoded formula resolution without LLM."""

    def test_irr(self) -> None:
        assert get_fallback_formula("calculate irr") == "=IRR(B2:B10)"

    def test_npv(self) -> None:
        assert get_fallback_formula("what is npv for these cashflows") == "=NPV(0.1,B2:B10)"

    def test_growth_rate(self) -> None:
        assert get_fallback_formula("compute growth rate please") == "=(B2-A2)/A2"

    def test_no_match(self) -> None:
        assert get_fallback_formula("random gibberish xyz qwerty") is None

    def test_empty(self) -> None:
        assert get_fallback_formula("") is None


# ──────────────────────────────────────────────────────────────────────
# Excel writer: write_formula, find_next_empty_cell
# ──────────────────────────────────────────────────────────────────────


class TestExcelWriter:
    """Round-trip a formula through openpyxl in a temp file."""

    def test_write_formula_round_trip(self, tmp_path: Path) -> None:
        path = tmp_path / "rt.xlsx"
        ok = write_formula(str(path), "Sheet1", "B1", "=SUM(A1:A5)")
        assert ok is True
        wb = load_workbook(path)
        assert wb["Sheet1"]["B1"].value == "=SUM(A1:A5)"

    def test_write_formula_rejects_bad_cell(self, tmp_path: Path) -> None:
        path = tmp_path / "bad.xlsx"
        with pytest.raises(ValueError):
            write_formula(str(path), "Sheet1", "1A", "=SUM(A1:A5)")

    def test_write_formula_rejects_bad_formula(self, tmp_path: Path) -> None:
        path = tmp_path / "bad.xlsx"
        with pytest.raises(ValueError):
            write_formula(str(path), "Sheet1", "B1", "SUM(A1:A5)")

    def test_find_next_empty_cell_preferred(self, tmp_path: Path) -> None:
        path = tmp_path / "find.xlsx"
        write_formula(str(path), "Sheet1", "B1", "=1+1")
        wb = load_workbook(path)
        ws = wb["Sheet1"]
        ws["B2"] = "x"
        ws["B3"] = "y"
        result = find_next_empty_cell(ws, preferred_col="B")
        assert result == "B4"


# ──────────────────────────────────────────────────────────────────────
# create_table: depreciation
# ──────────────────────────────────────────────────────────────────────


class TestCreateTable:
    """Structural assertions on generated tables."""

    def test_depreciation_schedule(self, tmp_path: Path) -> None:
        path = tmp_path / "dep.xlsx"
        ok = create_table(
            str(path),
            "Dep",
            {
                "type": "depreciation",
                "asset_value": 2_000_000,
                "rate": 0.25,
                "years": 5,
                "start_row": 1,
                "start_col": 1,
            },
        )
        assert ok is True
        wb = load_workbook(path)
        ws = wb["Dep"]
        # Headers
        assert ws["A1"].value == "Year"
        assert ws["B1"].value == "Opening Value"
        assert ws["C1"].value == "Depreciation"
        assert ws["D1"].value == "Closing Value"
        # 5 data rows
        for r in range(2, 7):
            assert ws.cell(row=r, column=1).value == r - 1
        # Formulas (not hardcoded values) in Depreciation column
        for r in range(2, 7):
            val = ws.cell(row=r, column=3).value
            assert isinstance(val, str) and val.startswith("=")
        # Closing value for row 1 is a formula referencing B2
        assert isinstance(ws["D2"].value, str) and ws["D2"].value.startswith("=")

    def test_amortization_table(self, tmp_path: Path) -> None:
        path = tmp_path / "amort.xlsx"
        ok = create_table(
            str(path),
            "Amort",
            {
                "type": "amortization",
                "principal": 5_000_000,
                "annual_rate": 0.09,
                "tenure_months": 12,
            },
        )
        assert ok is True
        wb = load_workbook(path)
        ws = wb["Amort"]
        assert ws["A1"].value == "Month"
        assert "PMT" in str(ws["C2"].value)
        assert "IPMT" in str(ws["D2"].value)
        assert "PPMT" in str(ws["E2"].value)
        # 12 month rows
        assert ws.cell(row=13, column=1).value == 12

    def test_projection_table(self, tmp_path: Path) -> None:
        path = tmp_path / "proj.xlsx"
        ok = create_table(
            str(path),
            "Proj",
            {
                "type": "projection",
                "base_revenue": 5_000_000,
                "growth_rate": 0.15,
                "cost_ratio": 0.6,
                "years": 5,
            },
        )
        assert ok is True
        wb = load_workbook(path)
        ws = wb["Proj"]
        assert ws["A1"].value == "Year"
        assert ws["B1"].value == "Revenue"
        # EBITDA margin column should be a formula
        for r in range(2, 7):
            v = ws.cell(row=r, column=5).value
            assert isinstance(v, str) and v.startswith("=")


# ──────────────────────────────────────────────────────────────────────
# detect_task_type — keyword fallback path (offline)
# ──────────────────────────────────────────────────────────────────────


class TestDetectTaskType:
    """The keyword fallback runs without network, so we test it directly."""

    def test_table_keyword(self) -> None:
        from agents.ai_engine import _keyword_classify  # noqa: PLC0415

        assert _keyword_classify("create depreciation schedule")["type"] == "table"

    def test_formula_default(self) -> None:
        from agents.ai_engine import _keyword_classify  # noqa: PLC0415

        assert _keyword_classify("growth rate A2 B2")["type"] == "formula"

    def test_chart_keyword(self) -> None:
        from agents.ai_engine import _keyword_classify  # noqa: PLC0415

        assert _keyword_classify("plot a chart of revenue")["type"] == "chart"


# ──────────────────────────────────────────────────────────────────────
# Finance-domain table_config validation
# ──────────────────────────────────────────────────────────────────────


class TestValidateTableConfig:
    """Domain-specific bounds enforced by validate_table_config."""

    def test_valid_depreciation_wdv(self) -> None:
        ok, _ = validate_table_config({
            "type": "depreciation",
            "asset_value": 2_000_000,
            "rate": 0.25,
            "years": 5,
        })
        assert ok is True

    def test_valid_depreciation_straight_line(self) -> None:
        ok, _ = validate_table_config({
            "type": "depreciation",
            "asset_value": 2_000_000,
            "rate": 0.25,  # ignored for straight-line, still validated as benign
            "years": 5,
            "method": "straight_line",
            "salvage_value": 100_000,
        })
        assert ok is True

    def test_unknown_method_rejected(self) -> None:
        ok, err = validate_table_config({
            "type": "depreciation",
            "asset_value": 1_000_000, "rate": 0.1, "years": 5,
            "method": "double_declining",
        })
        assert ok is False
        assert "method" in err

    def test_negative_asset_value_rejected(self) -> None:
        ok, err = validate_table_config({
            "type": "depreciation",
            "asset_value": -1, "rate": 0.25, "years": 5,
        })
        assert ok is False
        assert "asset_value" in err

    def test_rate_above_one_rejected_for_wdv(self) -> None:
        ok, err = validate_table_config({
            "type": "depreciation",
            "asset_value": 100_000, "rate": 2.5, "years": 5,
        })
        assert ok is False
        assert "rate" in err

    def test_salvage_above_asset_rejected(self) -> None:
        ok, err = validate_table_config({
            "type": "depreciation",
            "asset_value": 100_000, "rate": 0.25, "years": 5,
            "method": "straight_line",
            "salvage_value": 200_000,
        })
        assert ok is False
        assert "salvage_value" in err

    def test_amortization_zero_rate_allowed(self) -> None:
        ok, _ = validate_table_config({
            "type": "amortization",
            "principal": 1_000_000, "annual_rate": 0.0, "tenure_months": 60,
        })
        assert ok is True

    def test_amortization_tenure_too_long(self) -> None:
        ok, err = validate_table_config({
            "type": "amortization",
            "principal": 1_000_000, "annual_rate": 0.05, "tenure_months": 1000,
        })
        assert ok is False
        assert "tenure_months" in err

    def test_projection_growth_rate_fat_finger_rejected(self) -> None:
        # 50 (5000%) is the classic typo for 0.50 (50%).
        ok, err = validate_table_config({
            "type": "projection",
            "base_revenue": 5_000_000, "growth_rate": 50, "years": 5,
        })
        assert ok is False
        assert "growth_rate" in err

    def test_projection_negative_growth_allowed(self) -> None:
        ok, _ = validate_table_config({
            "type": "projection",
            "base_revenue": 5_000_000, "growth_rate": -0.1, "years": 5,
        })
        assert ok is True

    def test_projection_decline_to_zero_rejected(self) -> None:
        ok, err = validate_table_config({
            "type": "projection",
            "base_revenue": 5_000_000, "growth_rate": -1, "years": 5,
        })
        assert ok is False
        assert "growth_rate" in err

    def test_projection_invalid_growth_method(self) -> None:
        ok, err = validate_table_config({
            "type": "projection",
            "base_revenue": 5_000_000, "growth_rate": 0.1, "years": 5,
            "growth_method": "exponential",
        })
        assert ok is False
        assert "growth_method" in err

    def test_unknown_table_type_rejected(self) -> None:
        ok, err = validate_table_config({"type": "balance_sheet"})
        assert ok is False
        assert "table type" in err

    def test_create_table_rejects_invalid_config(self, tmp_path: Path) -> None:
        path = tmp_path / "bad.xlsx"
        with pytest.raises(ValueError, match="Invalid table config"):
            create_table(str(path), "X", {
                "type": "depreciation",
                "asset_value": -100, "rate": 0.25, "years": 5,
            })


# ──────────────────────────────────────────────────────────────────────
# Straight-line depreciation builder
# ──────────────────────────────────────────────────────────────────────


class TestStraightLineDepreciation:
    """The straight-line method writes a constant per-year amount with ROUND."""

    def test_constant_depreciation_amount(self, tmp_path: Path) -> None:
        path = tmp_path / "sl.xlsx"
        ok = create_table(str(path), "Dep", {
            "type": "depreciation",
            "asset_value": 1_000_000,
            "rate": 0.0,  # unused for SL
            "years": 5,
            "method": "straight_line",
            "salvage_value": 100_000,
            "start_row": 1,
            "start_col": 1,
        })
        assert ok is True
        wb = load_workbook(path)
        ws = wb["Dep"]
        # All five depreciation cells should be ROUND formulas referencing
        # the cost / salvage / life cells (no hardcoded numbers in the formula).
        for r in range(2, 7):
            v = ws.cell(row=r, column=3).value
            assert isinstance(v, str)
            assert v.startswith("=ROUND(")
            assert "$" in v  # absolute refs to assumption cells
        # Closing value drops by the same amount each year.
        # We can't evaluate Excel formulas in openpyxl, but we can verify
        # that all five depreciation formulas reference the same cost/salvage/life cells.
        formulas = [ws.cell(row=r, column=3).value for r in range(2, 7)]
        assert all(f == formulas[0] for f in formulas), \
            "straight-line depreciation should be identical across years"


# ──────────────────────────────────────────────────────────────────────
# Depreciation ROUND wrapper (WDV)
# ──────────────────────────────────────────────────────────────────────


class TestDepreciationRound:
    """WDV depreciation cells must be wrapped in ROUND for total integrity."""

    def test_wdv_uses_round(self, tmp_path: Path) -> None:
        path = tmp_path / "wdv.xlsx"
        create_table(str(path), "Dep", {
            "type": "depreciation",
            "asset_value": 2_000_000,
            "rate": 0.25,
            "years": 5,
            "start_row": 1,
            "start_col": 1,
        })
        wb = load_workbook(path)
        ws = wb["Dep"]
        for r in range(2, 7):
            v = ws.cell(row=r, column=3).value
            assert isinstance(v, str) and v.startswith("=ROUND(")


# ──────────────────────────────────────────────────────────────────────
# Linear projection growth
# ──────────────────────────────────────────────────────────────────────


class TestLinearProjection:
    """growth_method='linear' applies a constant increment, not compounding."""

    def test_linear_formula_shape(self, tmp_path: Path) -> None:
        path = tmp_path / "lin.xlsx"
        create_table(str(path), "Proj", {
            "type": "projection",
            "base_revenue": 5_000_000,
            "growth_rate": 0.15,
            "cost_ratio": 0.6,
            "years": 5,
            "growth_method": "linear",
        })
        wb = load_workbook(path)
        ws = wb["Proj"]
        # Row 1 = base. Row 2..5 should each reference base * (1 + r * (n-1)),
        # NOT the previous row's revenue.
        # The base ref is column G (assumption block) row 2 — i.e. $G$2.
        # Year 2 (sheet row 3) revenue formula should mention "$G$2" but NOT "B2".
        year2 = ws.cell(row=3, column=2).value
        assert isinstance(year2, str)
        assert "$" in year2  # absolute ref to assumption cell
        # Compounding form would reference "B2" (previous row of the same column).
        # Linear form references the base assumption cell with a multiplier.
        assert "*1)" in year2.replace(" ", "") or "*(1+" in year2
        # Year 5 should have a different multiplier from Year 2.
        year5 = ws.cell(row=6, column=2).value
        assert year2 != year5


# ──────────────────────────────────────────────────────────────────────
# preview_table — in-memory dry-run output
# ──────────────────────────────────────────────────────────────────────


class TestPreviewTable:
    """preview_table builds in memory and never touches disk."""

    def test_depreciation_preview_no_file_written(self, tmp_path: Path) -> None:
        # The path doesn't exist; preview should succeed without creating it.
        cfg = {
            "type": "depreciation",
            "asset_value": 1_000_000, "rate": 0.2, "years": 3,
            "start_row": 1, "start_col": 1,
        }
        preview = preview_table(cfg, sheet_name="Dep")
        assert isinstance(preview, list)
        assert len(preview) > 0
        # Headers must be present.
        cells = {p["cell"]: p["value"] for p in preview}
        assert cells["A1"] == "Year"
        assert cells["B1"] == "Opening Value"
        assert cells["C1"] == "Depreciation"
        assert cells["D1"] == "Closing Value"
        # Three data rows produce ROUND formulas in column C.
        for r in range(2, 5):
            v = cells[f"C{r}"]
            assert isinstance(v, str) and v.startswith("=ROUND(")
        # Confirm no file was created.
        assert not (tmp_path / "dep.xlsx").exists()

    def test_projection_preview_returns_all_cells(self) -> None:
        cfg = {
            "type": "projection",
            "base_revenue": 1_000_000, "growth_rate": 0.10,
            "cost_ratio": 0.5, "years": 3,
            "start_row": 1, "start_col": 1,
        }
        preview = preview_table(cfg, sheet_name="Proj")
        sheets = {p["sheet"] for p in preview}
        assert sheets == {"Proj"}
        # Header row + 3 data rows × 5 columns + assumption block ≈ 20+ cells.
        assert len(preview) >= 15


# ──────────────────────────────────────────────────────────────────────
# Dry-run via execute_excel_task — orchestrator integration
# ──────────────────────────────────────────────────────────────────────


class TestDryRunOrchestrator:
    """execute_excel_task with dry_run=True must not write or audit-log.

    These tests stub the LLM classifier with the deterministic offline keyword
    classifier so they verify dry-run behaviour, not LLM responses.
    """

    def test_formula_dry_run_does_not_write(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from agents import task_executor  # noqa: PLC0415
        from utils import audit  # noqa: PLC0415

        # Force offline classifier; LLM is not consulted for this test.
        monkeypatch.setattr(task_executor, "detect_task_type", task_executor._keyword_classify)

        # Redirect audit writes to tmp_path so we can verify nothing was written.
        monkeypatch.setattr(audit.CONFIG, "log_dir", tmp_path)
        audit.init_session()

        path = tmp_path / "wb.xlsx"
        result = task_executor.execute_excel_task(
            "calculate IRR for the cashflows",
            filepath=path,
            sheet_name="Sheet1",
            dry_run=True,
        )
        assert result["success"] is True, f"expected success, got {result['error']}: {result['message']}"
        assert result["dry_run"] is True
        assert result["preview"] is not None
        assert len(result["preview"]) == 1
        assert result["preview"][0]["value"].startswith("=IRR")
        assert result["cell_written"] is None
        # No workbook should have been created.
        assert not path.exists()
        # The dry-run path does not emit a formula_written event, but other
        # session-level events may exist. What we care about: the audit log
        # contains no `formula_written` or `table_built` events.
        audit_dir = tmp_path / "audit"
        if audit_dir.exists():
            for f in audit_dir.glob("*.jsonl"):
                content = f.read_text(encoding="utf-8")
                assert "formula_written" not in content
                assert "table_built" not in content

    def test_table_dry_run_does_not_write(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        from agents import task_executor  # noqa: PLC0415
        from utils import audit  # noqa: PLC0415

        # Force offline classifier and offline heuristic so this test does not
        # depend on LLM availability or response stability.
        monkeypatch.setattr(task_executor, "detect_task_type", task_executor._keyword_classify)
        monkeypatch.setattr(
            task_executor,
            "extract_table_config",
            lambda instr: task_executor._heuristic_table_config(instr, strict=False),
        )

        monkeypatch.setattr(audit.CONFIG, "log_dir", tmp_path)
        audit.init_session()

        path = tmp_path / "wb.xlsx"
        result = task_executor.execute_excel_task(
            "create a 5-year depreciation schedule for 20 lakh at 25%",
            filepath=path,
            sheet_name="Dep",
            dry_run=True,
        )
        assert result["success"] is True, f"expected success, got {result['error']}: {result['message']}"
        assert result["dry_run"] is True
        assert result["preview"] is not None
        assert len(result["preview"]) > 5
        assert result["table_created"] is False
        assert not path.exists()
        audit_dir = tmp_path / "audit"
        if audit_dir.exists():
            for f in audit_dir.glob("*.jsonl"):
                content = f.read_text(encoding="utf-8")
                assert "formula_written" not in content
                assert "table_built" not in content


# ──────────────────────────────────────────────────────────────────────
# Strict-mode behaviour of the offline table heuristic
# ──────────────────────────────────────────────────────────────────────


class TestStrictModeHeuristic:
    """In strict mode the heuristic must refuse instructions it cannot identify."""

    def test_recognized_depreciation_returns_config(self) -> None:
        from agents.ai_engine import _heuristic_table_config  # noqa: PLC0415

        cfg = _heuristic_table_config(
            "create a 5-year depreciation schedule for 20 lakh at 25%",
            strict=True,
        )
        assert cfg is not None
        assert cfg["type"] == "depreciation"

    def test_recognized_amortization_returns_config(self) -> None:
        from agents.ai_engine import _heuristic_table_config  # noqa: PLC0415

        cfg = _heuristic_table_config(
            "amortization for 50 lakh loan at 9% for 10 years",
            strict=True,
        )
        assert cfg is not None
        assert cfg["type"] == "amortization"

    def test_recognized_projection_returns_config(self) -> None:
        from agents.ai_engine import _heuristic_table_config  # noqa: PLC0415

        cfg = _heuristic_table_config(
            "5-year revenue projection at 50 lakh growth 15%",
            strict=True,
        )
        assert cfg is not None
        assert cfg["type"] == "projection"

    def test_unrecognized_returns_none_in_strict(self) -> None:
        from agents.ai_engine import _heuristic_table_config  # noqa: PLC0415

        assert _heuristic_table_config("build me a generic plan", strict=True) is None

    def test_unrecognized_returns_default_in_lax(self) -> None:
        from agents.ai_engine import _heuristic_table_config  # noqa: PLC0415

        # Legacy behaviour: defaults to depreciation when nothing else matches.
        cfg = _heuristic_table_config("build me a generic plan", strict=False)
        assert cfg is not None
        assert cfg["type"] == "depreciation"


# NOTE: audit-log tests live in tests/test_audit.py — they exercise the
# Step 6 schema (hash chain, session IDs, daily files, verification, query).


# ──────────────────────────────────────────────────────────────────────
# text_to_formula — live LLM (skipped when key absent)
# ──────────────────────────────────────────────────────────────────────


@NEEDS_MISTRAL
class TestTextToFormulaLive:
    """Live tests that hit Mistral's API; auto-skipped without a key."""

    def test_growth_rate(self) -> None:
        from agents.ai_engine import text_to_formula  # noqa: PLC0415

        formula = text_to_formula("growth rate between A2 and B2", context={})
        assert formula.startswith("=")
        assert "B2" in formula and "A2" in formula

    def test_cagr(self) -> None:
        from agents.ai_engine import text_to_formula  # noqa: PLC0415

        formula = text_to_formula("CAGR from A2 to B2 over 5 years", context={})
        assert formula.startswith("=")
        assert "^" in formula and "(1/5)" in formula.replace(" ", "")

    def test_irr(self) -> None:
        from agents.ai_engine import text_to_formula  # noqa: PLC0415

        formula = text_to_formula("IRR for cashflows in B2:B8", context={})
        assert formula.startswith("=")
        assert "IRR" in formula
