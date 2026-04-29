"""Streamlit UI for the Excel Finance Automation Agent.

Run with: ``streamlit run app.py``

Layout:
- Sidebar: mode (strict / AI-assisted), API key, model, dry-run toggle,
  recent activity expander.
- Main panel: two tabs — "Run" (workflow) and "Audit Log" (session events).
- Footer: current mode + formulas-written counter.

The Mistral API key is held in ``st.session_state`` only — never written
back to ``.env``. CONFIG is mutated in-process via :func:`update_config`
so the orchestrator picks up sidebar changes per session.
"""

from __future__ import annotations

import csv
import io
import os
import tempfile
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from agents.task_executor import execute_excel_task
from config import CONFIG, update_config
from utils import audit


# ──────────────────────────────────────────────────────────────────────
# Page config — must be the first Streamlit call.
# ──────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Excel Finance Agent",
    page_icon="💹",
    layout="wide",
)

MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB
MODEL_OPTIONS = ["mistral-small-latest", "mistral-medium-latest", "mistral-large-latest"]
MODE_AI = "AI-assisted (Mistral)"
MODE_STRICT = "Strict (offline only)"


# ──────────────────────────────────────────────────────────────────────
# Session state
# ──────────────────────────────────────────────────────────────────────


def _init_state() -> None:
    """Seed all session-state keys we depend on. Idempotent — safe to call every rerun."""
    fresh_session = "session_id" not in st.session_state
    if fresh_session:
        st.session_state.session_id = str(uuid.uuid4())
        st.session_state.session_start = datetime.now(timezone.utc).isoformat()
        # Tell the audit module to use this session ID for chain/seq tracking.
        audit.init_session(st.session_state.session_id)

    if "api_key" not in st.session_state:
        # Seed from .env on first load; the user can clear or replace this.
        st.session_state.api_key = CONFIG.mistral_api_key
    if "mode" not in st.session_state:
        # Default to strict if no key is configured anywhere.
        st.session_state.mode = MODE_AI if CONFIG.mistral_api_key else MODE_STRICT
    if "dry_run" not in st.session_state:
        st.session_state.dry_run = True  # safe default
    if "model" not in st.session_state:
        st.session_state.model = CONFIG.mistral_model or MODEL_OPTIONS[0]
    if "workbook_path" not in st.session_state:
        st.session_state.workbook_path = None
    if "uploaded_name" not in st.session_state:
        st.session_state.uploaded_name = None
    if "uploaded_size" not in st.session_state:
        st.session_state.uploaded_size = 0
    if "uploaded_hash" not in st.session_state:
        st.session_state.uploaded_hash = None
    if "selected_sheet" not in st.session_state:
        st.session_state.selected_sheet = None
    if "audit_events" not in st.session_state:
        st.session_state.audit_events = []
    if "formulas_written" not in st.session_state:
        st.session_state.formulas_written = 0
    if "tables_built" not in st.session_state:
        st.session_state.tables_built = 0
    if "pending_dry_run" not in st.session_state:
        st.session_state.pending_dry_run = None

    if fresh_session:
        # Emit session_started AFTER all session_state keys are initialised so the
        # mode field is correct when the audit event is recorded.
        audit.record_session_started(mode=st.session_state.mode, surface="streamlit")


def _apply_session_to_config() -> None:
    """Push session-state values into CONFIG so the orchestrator sees them.

    In strict mode the API key is forced empty so an accidental LLM call
    raises immediately rather than silently using a leftover key.
    """
    is_strict = (st.session_state.mode == MODE_STRICT)
    update_config(
        strict_mode=is_strict,
        mistral_api_key=st.session_state.api_key if not is_strict else "",
        mistral_model=st.session_state.model,
    )


# ──────────────────────────────────────────────────────────────────────
# File helpers
# ──────────────────────────────────────────────────────────────────────


def _save_uploaded(uploaded) -> Path:
    """Persist an uploaded UploadedFile to a temp .xlsx path and return it."""
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    path = Path(tmp_path)
    with path.open("wb") as f:
        f.write(uploaded.getbuffer())
    return path


def _list_sheets(path: Path) -> list[str]:
    """Return sheet names from *path*, or ``[]`` on failure."""
    try:
        return load_workbook(path, read_only=True).sheetnames
    except Exception:  # noqa: BLE001
        return []


def _read_sheet_preview(path: Path, sheet_name: str, n_rows: int = 20) -> pd.DataFrame:
    """Read up to *n_rows* rows of *sheet_name* showing formula text where present.

    Uses ``data_only=False`` so freshly-written formulas appear as their
    formula string rather than ``None`` (openpyxl does not evaluate
    formulas; cached values only exist for sheets previously opened in
    Excel/LibreOffice).
    """
    try:
        wb = load_workbook(path, data_only=False)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(max_row=n_rows, values_only=True))
    except Exception as exc:  # noqa: BLE001
        return pd.DataFrame({"error": [f"Could not read sheet: {exc}"]})

    if not rows:
        return pd.DataFrame()

    # Treat row 1 as headers when it's all strings/blanks.
    if all(isinstance(c, str) or c is None for c in rows[0]):
        headers = [c if c is not None else "" for c in rows[0]]
        return pd.DataFrame(rows[1:], columns=headers)
    return pd.DataFrame(rows)


# ──────────────────────────────────────────────────────────────────────
# Audit-event mirror (session_state copy of utils/audit.py disk writes)
# ──────────────────────────────────────────────────────────────────────


def _record_session_event(result: dict[str, Any], sheet_name: str) -> None:
    """Mirror what the orchestrator just wrote to disk into session_state."""
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.") + "000Z"
    file_name = Path(st.session_state.uploaded_name or "workbook.xlsx").name
    source = result.get("source") or "unknown"

    if result.get("formula") and result.get("cell_written"):
        st.session_state.audit_events.append({
            "ts": ts,
            "event": "formula_written",
            "file": file_name,
            "sheet": sheet_name,
            "cell": result["cell_written"],
            "formula": result["formula"],
            "source": source,
        })
        st.session_state.formulas_written += 1
    elif result.get("table_created"):
        st.session_state.audit_events.append({
            "ts": ts,
            "event": "table_built",
            "file": file_name,
            "sheet": sheet_name,
            "table_type": result.get("task_type") or "unknown",
            "source": source,
        })
        st.session_state.tables_built += 1


# ──────────────────────────────────────────────────────────────────────
# Sidebar
# ──────────────────────────────────────────────────────────────────────


def _render_sidebar() -> None:
    with st.sidebar:
        st.header("⚙️ Settings")

        # -- Mode toggle --------------------------------------------------
        st.subheader("Mode")
        mode = st.radio(
            "Operation mode",
            options=[MODE_AI, MODE_STRICT],
            index=0 if st.session_state.mode == MODE_AI else 1,
            label_visibility="collapsed",
            key="mode_radio",
        )
        if mode != st.session_state.mode:
            audit.record_mode_changed(from_mode=st.session_state.mode, to_mode=mode)
            st.session_state.mode = mode
            st.session_state.pending_dry_run = None  # invalidate stale preview
        is_strict = (st.session_state.mode == MODE_STRICT)
        if is_strict:
            st.info(
                "🔒 **Strict mode.** No instruction or data leaves the machine. "
                "Only hardcoded formulas and the offline parser are used; "
                "instructions the offline path cannot resolve fail loudly."
            )

        # -- API key -----------------------------------------------------
        st.subheader("Mistral API key")
        if is_strict:
            st.text_input(
                "API key",
                value="",
                disabled=True,
                type="password",
                label_visibility="collapsed",
                placeholder="Disabled in strict mode",
            )
        else:
            key_input = st.text_input(
                "API key",
                value=st.session_state.api_key,
                type="password",
                label_visibility="collapsed",
                placeholder="paste key here",
                key="api_key_input",
            )
            cols = st.columns([1, 1])
            with cols[0]:
                if st.button("Save to session", use_container_width=True):
                    if key_input and key_input != st.session_state.api_key:
                        audit.record_api_key_loaded()
                    st.session_state.api_key = key_input
                    st.rerun()
            with cols[1]:
                if st.button("Clear", use_container_width=True):
                    if st.session_state.api_key:
                        audit.record_api_key_cleared()
                    st.session_state.api_key = ""
                    st.rerun()

            if st.session_state.api_key:
                st.markdown(":green[✓] Key loaded for this session")
            else:
                st.markdown(":red[✗] No key — paste one above")

        # -- Model selector (AI-assisted only) --------------------------
        if not is_strict:
            st.subheader("Model")
            model = st.selectbox(
                "Mistral model",
                options=MODEL_OPTIONS,
                index=MODEL_OPTIONS.index(st.session_state.model)
                if st.session_state.model in MODEL_OPTIONS else 0,
                label_visibility="collapsed",
            )
            if model != st.session_state.model:
                audit.record_model_selected(model=model)
                st.session_state.model = model
            st.caption("Small is free tier; medium and large are paid.")

        # -- Dry run -----------------------------------------------------
        st.subheader("Safety")
        dry_run = st.checkbox(
            "Dry run (preview before writing)",
            value=st.session_state.dry_run,
            help=(
                "When on, the agent shows the formulas it would write and "
                "waits for explicit confirmation before modifying the file."
            ),
        )
        if dry_run != st.session_state.dry_run:
            st.session_state.dry_run = dry_run
            st.session_state.pending_dry_run = None

        # -- Recent activity --------------------------------------------
        st.subheader("Recent activity")
        with st.expander(
            f"Last 5 events ({len(st.session_state.audit_events)} total)",
            expanded=False,
        ):
            recent = list(reversed(st.session_state.audit_events))[:5]
            if not recent:
                st.caption("No events yet this session.")
            else:
                for e in recent:
                    src = e.get("source", "?")
                    src_emoji = "🤖" if src == "llm" else "📋"
                    if e.get("event") == "formula_written":
                        st.caption(
                            f"{src_emoji} `{e['sheet']}!{e['cell']}` ← "
                            f"`{e['formula']}`"
                        )
                    elif e.get("event") == "table_built":
                        st.caption(
                            f"{src_emoji} `{e['sheet']}` — {e['table_type']} table"
                        )


# ──────────────────────────────────────────────────────────────────────
# Run tab
# ──────────────────────────────────────────────────────────────────────


def _render_run_tab() -> None:
    # 1. Upload
    st.subheader("1. Upload your workbook")
    uploaded = st.file_uploader(
        "xlsx file (max 50 MB)",
        type=["xlsx"],
        accept_multiple_files=False,
        key="file_uploader",
    )
    if uploaded is not None:
        if uploaded.size > MAX_FILE_SIZE_BYTES:
            st.error(
                f"File too large: {uploaded.size / 1_048_576:.1f} MB. "
                f"Maximum is {MAX_FILE_SIZE_BYTES / 1_048_576:.0f} MB."
            )
            return
        if not uploaded.name.lower().endswith(".xlsx"):
            st.error(f"Only .xlsx files are supported. Got: {uploaded.name}")
            return
        if uploaded.name != st.session_state.uploaded_name:
            saved = _save_uploaded(uploaded)
            file_hash = audit.file_sha256(saved)
            audit.record_file_uploaded(
                file=uploaded.name,
                size_bytes=uploaded.size,
                file_hash=file_hash,
            )
            st.session_state.workbook_path = saved
            st.session_state.uploaded_name = uploaded.name
            st.session_state.uploaded_size = uploaded.size
            st.session_state.uploaded_hash = file_hash
            st.session_state.pending_dry_run = None
            st.session_state.selected_sheet = None  # force re-record on first pick
            st.success(f"Loaded **{uploaded.name}** ({uploaded.size / 1024:.1f} KB)")

    if st.session_state.workbook_path is None:
        st.info("Upload a workbook to get started.")
        return

    path: Path = st.session_state.workbook_path
    sheet_names = _list_sheets(path)
    if not sheet_names:
        st.error("Could not read sheet names from the uploaded file.")
        return

    # 2. Sheet picker + preview
    st.subheader("2. Choose a sheet and review")
    sheet_name = st.selectbox("Sheet", options=sheet_names, key="sheet_selector")
    if sheet_name != st.session_state.selected_sheet:
        audit.record_sheet_selected(file=st.session_state.uploaded_name, sheet=sheet_name)
        st.session_state.selected_sheet = sheet_name

    df = _read_sheet_preview(path, sheet_name, n_rows=20)
    st.caption(f"First 20 rows of **{sheet_name}** (formulas shown as text):")
    st.dataframe(df, use_container_width=True, hide_index=True)

    # 3. Instruction
    st.subheader("3. What do you want done?")
    instruction = st.text_area(
        "Instruction",
        height=120,
        placeholder=(
            "Examples:\n"
            "  • Calculate the average of column B\n"
            "  • Build a 5-year projection at 50 lakh with 15% growth\n"
            "  • Create a 5-year WDV depreciation schedule for 20 lakh at 25%\n"
            "  • Build amortization for 50 lakh loan at 9% for 10 years"
        ),
        label_visibility="collapsed",
        key="instruction_input",
    )
    target_cell = st.text_input(
        "Target cell (optional)",
        value="",
        placeholder="e.g. C5",
        help="Leave blank to write to the next empty cell.",
        key="target_cell_input",
    )

    # Pre-flight check
    is_strict = (st.session_state.mode == MODE_STRICT)
    if not is_strict and not st.session_state.api_key:
        st.error(
            "⚠ No API key configured. Either paste a Mistral key in the "
            "**sidebar**, or switch to **Strict mode**."
        )
        return

    # 4. Generate
    st.subheader("4. Generate")
    generate = st.button(
        "▶ Generate",
        type="primary",
        use_container_width=True,
        key="generate_btn",
    )

    if generate:
        if not instruction.strip():
            st.warning("Please enter an instruction first.")
            return

        # Reset any stale pending preview before kicking off a new run.
        st.session_state.pending_dry_run = None
        _apply_session_to_config()

        with st.status("Working…", expanded=False) as status:
            try:
                result = execute_excel_task(
                    instruction=instruction.strip(),
                    filepath=path,
                    sheet_name=sheet_name,
                    cell=target_cell.strip() or None,
                    dry_run=st.session_state.dry_run,
                )
                status.update(label="Done", state="complete")
            except Exception as exc:  # noqa: BLE001
                status.update(label="Failed", state="error")
                st.error(f"Unexpected error: {type(exc).__name__}: {exc}")
                return

        if not result.get("success"):
            err = result.get("error") or "Failed"
            msg = result.get("message") or ""
            st.error(f"**{err}** — {msg}")
            return

        if result.get("dry_run"):
            st.session_state.pending_dry_run = {
                "result": result,
                "instruction": instruction.strip(),
                "sheet": sheet_name,
                "cell": target_cell.strip() or None,
            }
            st.success("Preview generated. Review below and confirm.")
        else:
            _record_session_event(result, sheet_name)
            st.success(result.get("message") or "Done.")
            _render_post_write_section(path, sheet_name)

    # If a pending dry-run preview exists (from this rerun or the previous), render it.
    if st.session_state.pending_dry_run is not None:
        _render_dry_run_preview()


def _render_dry_run_preview() -> None:
    """Show the pending dry-run preview with Confirm / Cancel actions."""
    pending = st.session_state.pending_dry_run
    if pending is None:
        return

    result = pending["result"]
    preview = result.get("preview") or []

    st.subheader("Preview")
    st.caption(result.get("message") or "")

    if not preview:
        st.warning("Nothing to preview.")
        return

    df = pd.DataFrame(preview)
    st.dataframe(df, use_container_width=True, hide_index=True)

    cols = st.columns([1, 1, 4])
    with cols[0]:
        confirm = st.button(
            "✓ Confirm and write",
            type="primary",
            use_container_width=True,
            key="confirm_write_btn",
        )
    with cols[1]:
        cancel = st.button(
            "✗ Cancel",
            use_container_width=True,
            key="cancel_dry_run_btn",
        )

    if cancel:
        audit.record_dry_run_cancelled(sheet=pending["sheet"])
        st.session_state.pending_dry_run = None
        st.rerun()

    if confirm:
        audit.record_dry_run_confirmed(sheet=pending["sheet"])
        _apply_session_to_config()
        path = st.session_state.workbook_path
        with st.status("Writing…", expanded=False) as status:
            try:
                final = execute_excel_task(
                    instruction=pending["instruction"],
                    filepath=path,
                    sheet_name=pending["sheet"],
                    cell=pending["cell"],
                    dry_run=False,
                )
                status.update(label="Written", state="complete")
            except Exception as exc:  # noqa: BLE001
                status.update(label="Failed", state="error")
                st.error(f"Unexpected error: {type(exc).__name__}: {exc}")
                return

        if final.get("success"):
            _record_session_event(final, pending["sheet"])
            st.session_state.pending_dry_run = None
            st.success(final.get("message") or "Done.")
            _render_post_write_section(path, pending["sheet"])
        else:
            st.error(f"{final.get('error')}: {final.get('message')}")


def _render_post_write_section(path: Path, sheet_name: str) -> None:
    """Render the updated-preview + download button after a successful write."""
    st.subheader("Updated workbook")
    df = _read_sheet_preview(path, sheet_name, n_rows=20)
    st.caption(f"First 20 rows of **{sheet_name}** after the write:")
    st.dataframe(df, use_container_width=True, hide_index=True)

    if path and path.exists():
        with path.open("rb") as f:
            data = f.read()
        if st.download_button(
            label="📥 Download updated workbook",
            data=data,
            file_name=st.session_state.uploaded_name or "workbook.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_btn",
        ):
            audit.record_file_downloaded(file=st.session_state.uploaded_name or "workbook.xlsx")


# ──────────────────────────────────────────────────────────────────────
# Audit Log tab
# ──────────────────────────────────────────────────────────────────────


def _flatten_event(ev: dict[str, Any]) -> dict[str, Any]:
    """Flatten an audit event for tabular display (payload keys hoisted to top level)."""
    flat = {
        "ts": ev.get("ts"),
        "seq": ev.get("seq"),
        "session_id": ev.get("session_id"),
        "event_type": ev.get("event_type"),
    }
    payload = ev.get("payload") or {}
    for k, v in payload.items():
        # Truncate long instruction text in the table view; full text remains
        # in the underlying JSONL.
        if k == "instruction" and isinstance(v, str) and len(v) > 80:
            flat[k] = v[:80] + "…"
        else:
            flat[k] = v
    return flat


def _render_audit_tab() -> None:
    st.subheader("Audit log")
    st.caption(
        "Every action the agent took. Cell values from the workbook are NEVER "
        "recorded — only formulas (what was written), target locations, the "
        "user's instruction text, and file SHA-256 hashes."
    )

    # Date picker — defaults to today (UTC).
    today = datetime.now(timezone.utc).date()
    cols = st.columns([2, 2, 2, 2])
    with cols[0]:
        picked = st.date_input("Date (UTC)", value=today, key="audit_date_picker")
    with cols[1]:
        scope = st.radio(
            "Scope",
            options=["This session", "All sessions"],
            index=0,
            horizontal=True,
            key="audit_scope_radio",
        )
    with cols[2]:
        verify_clicked = st.button(
            "🔍 Verify integrity",
            use_container_width=True,
            key="audit_verify_btn",
        )
    with cols[3]:
        refresh = st.button(
            "🔄 Refresh",
            use_container_width=True,
            key="audit_refresh_btn",
        )

    date_str = picked.strftime("%Y-%m-%d")

    # Verify integrity: run the chain check on the displayed day's file.
    if verify_clicked:
        breaks = audit.verify_audit_log(date_str)
        if not breaks:
            st.success(
                f"✓ Hash chain intact for **{date_str}**. "
                "Every event in every session links to the previous one with a valid SHA-256."
            )
        else:
            st.error(
                f"✗ {len(breaks)} chain break{'s' if len(breaks) != 1 else ''} "
                f"detected in **{date_str}**:"
            )
            st.dataframe(pd.DataFrame(breaks), use_container_width=True, hide_index=True)

    # Read events for the picked day.
    events = audit.read_audit_file(date_str)
    if scope == "This session":
        events = [e for e in events if e.get("session_id") == st.session_state.session_id]

    if not events:
        st.info(
            f"No events found for {date_str}"
            + (" in this session." if scope == "This session" else ".")
            + " Generate a formula or table on the **Run** tab to populate the log."
        )
        return

    # Filters
    flat = [_flatten_event(e) for e in events]
    df_full = pd.DataFrame(flat)

    fcols = st.columns([3, 2, 2])
    with fcols[0]:
        all_event_types = sorted({e.get("event_type", "") for e in events if e.get("event_type")})
        selected_types = st.multiselect(
            "Event type",
            options=all_event_types,
            default=[],
            placeholder="All types",
            key="audit_event_type_filter",
        )
    with fcols[1]:
        file_filter = st.text_input(
            "File name contains",
            value="",
            placeholder="e.g. deal_model",
            key="audit_file_filter",
        )
    with fcols[2]:
        source_filter = st.selectbox(
            "Source",
            options=["All", "llm", "offline_fallback", "offline_classifier"],
            index=0,
            key="audit_source_filter",
        )

    df = df_full
    if selected_types:
        df = df[df["event_type"].isin(selected_types)]
    if file_filter and "file" in df.columns:
        df = df[df["file"].fillna("").astype(str).str.contains(file_filter, case=False, na=False)]
    if source_filter != "All" and "source" in df.columns:
        df = df[df["source"] == source_filter]

    st.caption(f"{len(df)} of {len(df_full)} events shown")
    st.dataframe(df.iloc[::-1], use_container_width=True, hide_index=True)

    # CSV export of filtered rows.
    if not df.empty:
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        st.download_button(
            label="📥 Download filtered events (CSV)",
            data=csv_buffer.getvalue(),
            file_name=f"audit_{date_str}_{scope.replace(' ', '_').lower()}.csv",
            mime="text/csv",
            key="audit_csv_download",
        )


# ──────────────────────────────────────────────────────────────────────
# Footer
# ──────────────────────────────────────────────────────────────────────


def _render_footer() -> None:
    n = st.session_state.formulas_written
    t = st.session_state.tables_built
    parts = []
    if n:
        parts.append(f"{n} formula{'s' if n != 1 else ''}")
    if t:
        parts.append(f"{t} table{'s' if t != 1 else ''}")
    summary = ", ".join(parts) if parts else "0 formulas"
    st.markdown("---")
    st.caption(
        f"Running in **{st.session_state.mode}** · {summary} written this session."
    )


# ──────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────


def main() -> None:
    _init_state()
    _apply_session_to_config()

    st.title("💹 Excel Finance Agent")
    st.caption(
        "Plain-English finance instructions → Excel formulas and tables. "
        "Built for analysts who'd rather describe a depreciation schedule "
        "than wire it up cell by cell."
    )

    _render_sidebar()

    tab_run, tab_audit = st.tabs(["▶ Run", "📋 Audit Log"])
    with tab_run:
        _render_run_tab()
    with tab_audit:
        _render_audit_tab()

    _render_footer()


if __name__ == "__main__":
    main()
